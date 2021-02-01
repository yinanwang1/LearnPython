
import os
import openpyxl

phone_numbers = set()


def save():
    """
    将订单详情保存到excel
    :param order_info_list:
    :return:
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "开锁过的用户手机号"
    sheet.cell(row=1, column=1, value='用户手机号')
    numbers = list(phone_numbers)

    for row in range(0, len(numbers)):
        number = numbers[row]
        sheet.cell(row=row + 1, column=1, value=number)

    desk_top = os.path.join(os.path.expanduser("~"), 'Desktop')
    target_dir = os.path.abspath(desk_top)
    file_path = os.path.join(target_dir, "开锁过的用户手机号.xlsx")
    wb.save(file_path)


f = open("/Users/arthurwang/Desktop/2020-08-06/access.log", "r")
# f = open("/Users/arthurwang/Desktop/2020-08-06/access.log", "r")
lines = f.readlines()      #读取全部内容 ，并以列表方式返回
for line in lines:
    try:
        line.index('/dbike/bike/unlock')
        phone_index = line.index('phone=')
        phone_index += 6
        phone_index_end = phone_index + 11
        phone_number = line[phone_index:phone_index_end]
        phone_numbers.add(phone_number)
    except:
        pass


save()
