
# 查找ip地址，并统计
# https://kibana.youxu.cc/app/kibana#/discover?_g=(refreshInterval:(display:Off,pause:!f,value:0),time:(from:now-12h,mode:quick,to:now))&_a=(columns:!(type,message),filters:!(('$state':(store:appState),meta:(alias:!n,disabled:!f,index:AYL4WPDxV7HMEnaKCMRS,key:type,negate:!t,type:phrase,value:tothirdparty),query:(match:(type:(query:tothirdparty,type:phrase)))),('$state':(store:appState),meta:(alias:!n,disabled:!f,index:AYL4WPDxV7HMEnaKCMRS,key:message,negate:!f,type:phrase,value:%2Fuser%2Fuser%2Fverify),query:(match:(message:(query:%2Fuser%2Fuser%2Fverify,type:phrase))))),index:AYL4WPDxV7HMEnaKCMRS,interval:auto,query:(match_all:()),sort:!('@timestamp',desc))
# 查询/user/verify的最近12个小时的调用
# 根据ip地址和请求的身份证号，进行分组。计算同一个ip地址，不同身份证的数量。
# 大致可以判断，这个ip地址存在刷单的嫌疑。

import json
import os

import openpyxl

addTime = 'addTime'
latitude = 'latitude'
longitude = 'longitude'


def save(ipsDict: dict):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "注册的ip地址"
    sheet.cell(row=1, column=1, value='ip地址')
    sheet.cell(row=1, column=2, value='身份证号')
    sheet.cell(row=1, column=3, value='次数')

    keys = ipsDict.keys()
    row = 0
    for key in keys:
        identity_set = ipsDict.get(key)
        sheet.cell(row=row + 2, column=1, value=key)
        sheet.cell(row=row + 2, column=2, value=','.join(identity_set))
        sheet.cell(row=row + 2, column=3, value=len(identity_set))
        row = row + 1

    desk_top = os.path.join(os.path.expanduser("~"), 'Desktop')
    target_dir = os.path.abspath(desk_top)
    file_path = os.path.join(target_dir, "注册的ip地址.xlsx")
    wb.save(file_path)


def fetchIpAndIdentityNoSet(content: str) -> dict:
    model = json.loads(content)
    try:
        hits = model["hits"]["hits"]
        source_dict = dict()
        identity_no_key = "identity_no="
        identity_length = 18

        for hit in hits:
            source = hit["_source"]
            message = source["message"]
            index = message.find(" - - ")
            # print("message is {}".format(message))
            identity_no_index = message.find(identity_no_key)
            # print("index is {}, identity_no_index is {}".format(index, identity_no_index))
            if index != -1 and identity_no_index != -1:
                ip = message[:index]
                identity_no_index += len(identity_no_key)
                identity_no = message[identity_no_index:identity_no_index + identity_length]
                value = source_dict.get(ip)
                if value is None:
                    source_dict[ip] = {identity_no}
                else:
                    value.add(identity_no)
                    source_dict[ip] = value

        return source_dict

    except Exception as e:
        print(e)
        return None


def main():
    with open('/Users/arthurwang/Desktop/1.txt', 'r') as f:
        content = f.read()
        ip_and_identities = fetchIpAndIdentityNoSet(content)
        print(ip_and_identities)

        save(ip_and_identities)


if __name__ == "__main__":
    main()

