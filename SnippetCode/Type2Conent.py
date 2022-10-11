
# 从桌面的log日志文件中获取单车type为2的内容，并将结果保存桌面

import json
import os
import time

import openpyxl

battery_quantity = 'battery_quantity'
location = 'location'
battery_time = 'battery_time'


def save(batteryList: list):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "5710507070电池记录"
    sheet.cell(row=1, column=1, value='时间')
    sheet.cell(row=1, column=2, value='电量')
    sheet.cell(row=1, column=3, value='经纬度')

    for row in range(0, len(batteryList)):
        dispatcher = batteryList[row][battery_time]
        sheet.cell(row=row + 2, column=1, value=dispatcher)

        number = batteryList[row][battery_quantity]
        sheet.cell(row=row + 2, column=2, value=number)

        battery_location = batteryList[row][location]
        sheet.cell(row=row + 2, column=3, value=battery_location)

    desk_top = os.path.join(os.path.expanduser("~"), 'Desktop')
    target_dir = os.path.abspath(desk_top)
    file_path = os.path.join(target_dir, "5710507070电池记录.xlsx")
    wb.save(file_path)


def fetchType2(content: str) -> dict:
    index = content.find('send mq:{"type":2,"bikeNo":"5710507070"')
    if index == -1:
        return None

    message = 'send mq:'
    index = content.find(message)
    type2Content = content[index + len(message):]
    model = json.loads(type2Content)
    try:
        type2ContentStr = model["content"] if model["content"] else ""
        values = type2ContentStr.split(",")
        print(values)
        if len(values) != 6:
            return None

        time_temp = values[4] if values[4] else "0"
        time_local = time.localtime(int(time_temp))
        time_format = time.strftime("%Y-%m-%d %H:%M:%S", time_local)

        result = dict()
        result[battery_quantity] = values[1]
        result[location] = "{},{}".format(values[2], values[3])
        result[battery_time] = time_format

        return result

    except Exception as e:
        print(e)
        return None

def main():
    names = ['/Users/arthurwang/Desktop/info.2020-12-22 14.log',
             '/Users/arthurwang/Desktop/info.2020-12-22 15.log',
             '/Users/arthurwang/Desktop/info.2020-12-22 16.log',
             '/Users/arthurwang/Desktop/info.2020-12-22 17.log',
             '/Users/arthurwang/Desktop/info.2020-12-22 18.log']

    # 包含type=2的内容的model
    batteryList = list()

    for name in names:
        with open(name, "r") as fo:
            while True:
                line = fo.readline()
                if not line:
                    break
                content = fetchType2(line)
                if content:
                    batteryList.append(content)

    save(batteryList)


if __name__ == "__main__":
    main()
