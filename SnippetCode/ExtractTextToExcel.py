
# 从桌面的log日志文件中获取单车type为2的内容，并将结果保存桌面

import datetime
import json
import os
import time

import openpyxl

"""
获取一个json文件的内容，
将内容转化为 excel
包含：时间的转换
"""


def save(batteryList: list):
    print("111")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "经纬度轨迹"
    sheet.cell(row=1, column=1, value='距离')
    sheet.cell(row=1, column=2, value='骑行模式')
    sheet.cell(row=1, column=3, value='添加时间')
    sheet.cell(row=1, column=4, value='订单号')
    sheet.cell(row=1, column=5, value='单车ID')
    sheet.cell(row=1, column=6, value='经纬度')


    for row in range(0, len(batteryList)):
        source = batteryList[row]["_source"]

        mileage = source["currentMileage"]
        sheet.cell(row=row + 2, column=1, value=mileage)

        mode = source["driveMode"]
        sheet.cell(row=row + 2, column=2, value=mode)

        addTime = source["addTime"]
        chinaTime = changeUTCToLocalTime(addTime)
        sheet.cell(row=row + 2, column=3, value=chinaTime)

        orderId = source["orderId"]
        sheet.cell(row=row + 2, column=4, value=orderId)

        bikeId = source["bikeId"]
        sheet.cell(row=row + 2, column=5, value=bikeId)

        latitude = source["latitude"]
        longitude = source["longitude"]
        sheet.cell(row=row + 2, column=6, value="{},{}".format(longitude, latitude))

    desk_top = os.path.join(os.path.expanduser("~"), 'Desktop')
    target_dir = os.path.abspath(desk_top)
    file_path = os.path.join(target_dir, "经纬度轨迹.xlsx")
    wb.save(file_path)


def changeTimestampToLocalTime(addTime: str) -> str:
    try:
        time_temp = addTime if addTime else "0"
        time_local = time.localtime(int(time_temp))
        time_format = time.strftime("%Y-%m-%d %H:%M:%S", time_local)

        return time_format
    except Exception as e:
        print(e)
        return ""


def changeUTCToLocalTime(addTime: str) -> str:
    try:
        dateTimeP = datetime.datetime.strptime(addTime, '%Y-%m-%dT%H:%M:%S.%fZ')
        chinaTime = (dateTimeP + datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')

        return chinaTime
    except Exception as e:
        print(e)
        return ""


def main():
    file = '/Users/arthurwang/Desktop/gj.txt'

    with open(file, "r") as fo:
        content = fo.read()
        print("222")
        if content:
            print("333")
            # print(content)
            jsonObject = json.loads(content)
            print("jsonObject is {}".format(jsonObject))
            if isinstance(jsonObject, list):
                print("444")
                save(jsonObject)


if __name__ == "__main__":
    main()
