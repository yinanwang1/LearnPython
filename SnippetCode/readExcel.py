import os

import openpyxl

# 将两个Excel的表格，根据某一行相同，然后将两张合并在一起

desk_top = os.path.join(os.path.expanduser("~"), 'Desktop')
failedExcel = openpyxl.load_workbook(desk_top + "/0608共享单车失败数据-.xlsx")
failedSheet = failedExcel.active
print("failedSheet is {}".format(failedSheet))
print("A1 {}".format(failedSheet.max_row))

bikesExcel = openpyxl.load_workbook(desk_top + "/佛山车辆5644.xlsx")
bikesSheet = bikesExcel.active

print("bikesSheet is {}".format(bikesSheet))
print("A1 {}".format(bikesSheet.max_row))

failedSheet_max_column = failedSheet.max_column
bikesSheet_max_column = bikesSheet.max_column

for i in range(1, bikesSheet.max_column + 1):
    """
    第一行先添加好名称
    """
    failedSheet.cell(row=1, column=failedSheet_max_column + i, value=(bikesSheet.cell(row=1, column=i)).value)

failedSheet_max_row = failedSheet.max_row
bikesSheet_max_row = bikesSheet.max_row

for i in range(2, failedSheet_max_row + 1):
    print("failedSheet {}/{}".format(i, failedSheet_max_row))
    code = failedSheet.cell(row=i, column=3).value
    for j in range(2, bikesSheet_max_row + 1):
        codeInBike = bikesSheet.cell(row=j, column=8).value
        if code == codeInBike:
            # 找到对应的行，将数据写入
            for k in range(1, bikesSheet_max_column + 1):
                failedSheet.cell(row=i, column=failedSheet_max_column + k, value=bikesSheet.cell(row=j, column=k).value)
            continue

file_path = os.path.join(desk_top, "整合1.xlsx")
failedExcel.save(file_path)

print("DONE")

